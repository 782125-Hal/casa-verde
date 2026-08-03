"""Exportación de presupuestos a PDF y Excel (Fase 5 del diseño).

Dos documentos distintos, para dos momentos distintos:

- **Presupuesto** (PDF y Excel): lo que se presenta a socios o banco antes de
  ejecutar. Lleva el desglose de capas y las partidas por categoría.
- **Cierre de obra** (Excel): presupuestado vs. real por partida, con órdenes de
  cambio y KPIs. Es el documento de after-action, cuando la obra terminó.

Se elige **ReportLab** para el PDF: es Python puro y funciona en el cPanel del
proyecto, donde WeasyPrint no podría instalar cairo/pango. El precio es que el
layout se programa a mano en vez de heredarlo del HTML.

Ambos generadores devuelven ``bytes`` y no escriben a disco: la vista los sirve
directamente y no queda basura en el servidor.
"""

# El servidor corre Python 3.9: sin esto, una anotación como `Decimal | None`
# se evalúa al importar y truena con "unsupported operand type(s) for |".
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

VERDE = colors.HexColor('#1a7f37')
GRIS = colors.HexColor('#6e7781')
GRIS_CLARO = colors.HexColor('#f0f2f4')


def _pesos(valor) -> str:
    return f'${valor:,.2f}'


def _filas_por_categoria(presupuesto):
    """(etiqueta, partidas, subtotal) por categoría, en orden alfabético."""
    from presupuestos.views import _partidas_por_categoria

    return _partidas_por_categoria(presupuesto)


def _capas(presupuesto):
    """Las filas del desglose §2.4, con el costo del ROI señalado aparte."""
    filas = [
        ('Costo directo', presupuesto.subtotal_directo),
        (f'+ Indirectos ({presupuesto.pct_indirectos}%)', presupuesto.monto_indirectos),
        (f'+ Contingencia ({presupuesto.pct_contingencia}%)', presupuesto.monto_contingencia),
        (f'+ Utilidad ({presupuesto.pct_utilidad}%)', presupuesto.monto_utilidad),
    ]
    if presupuesto.aplica_iva:
        filas.append((f'+ IVA ({presupuesto.pct_iva}%)', presupuesto.monto_iva))
    return filas


# ---------------------------------------------------------------------------
# PDF del presupuesto
# ---------------------------------------------------------------------------

def presupuesto_a_pdf(presupuesto) -> bytes:
    """Presupuesto presentable: encabezado, partidas por categoría y capas."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
        topMargin=1.6 * cm, bottomMargin=1.6 * cm,
        title=f'Presupuesto — {presupuesto.nombre}',
    )
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle('titulo', parent=estilos['Heading1'], fontSize=16,
                            textColor=VERDE, spaceAfter=2)
    sub = ParagraphStyle('sub', parent=estilos['Normal'], fontSize=9, textColor=GRIS)
    celda = ParagraphStyle('celda', parent=estilos['Normal'], fontSize=8, leading=10)

    elementos = [Paragraph(presupuesto.nombre, titulo)]

    encabezado = [presupuesto.get_tipo_obra_display(), presupuesto.get_estado_display()]
    if presupuesto.propiedad:
        encabezado.append(presupuesto.propiedad.titulo)
    if presupuesto.area_m2:
        encabezado.append(f'{presupuesto.area_m2} m²')
    elementos += [Paragraph(' · '.join(encabezado), sub), Spacer(1, 0.6 * cm)]

    # --- Partidas, agrupadas por categoría --------------------------------
    datos = [['Concepto', 'Unidad', 'Cantidad', 'P.U.', 'Importe']]
    estilo_filas = [
        ('BACKGROUND', (0, 0), (-1, 0), VERDE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d0d7de')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]
    fila = 1
    for grupo in _filas_por_categoria(presupuesto):
        datos.append([Paragraph(f'<b>{grupo["etiqueta"].upper()}</b>', celda),
                      '', '', '', _pesos(grupo['subtotal'])])
        estilo_filas.append(('BACKGROUND', (0, fila), (-1, fila), GRIS_CLARO))
        fila += 1
        for partida in grupo['partidas']:
            datos.append([
                Paragraph(partida.descripcion, celda),
                partida.get_unidad_display(),
                f'{partida.cantidad:,.2f}',
                _pesos(partida.pu),
                _pesos(partida.importe),
            ])
            fila += 1

    if fila == 1:   # solo el encabezado: no hay partidas
        datos.append([Paragraph('<i>Sin partidas capturadas</i>', celda), '', '', '', ''])

    tabla = Table(datos, colWidths=[8.2 * cm, 1.9 * cm, 2.2 * cm, 2.5 * cm, 2.7 * cm],
                  repeatRows=1)
    tabla.setStyle(TableStyle(estilo_filas))
    elementos += [tabla, Spacer(1, 0.7 * cm)]

    # --- Capas + el costo que alimenta el ROI ------------------------------
    capas = [[etiqueta, _pesos(monto)] for etiqueta, monto in _capas(presupuesto)]
    capas.append(['TOTAL (comercial)', _pesos(presupuesto.total)])
    capas.append(['Costo que alimenta el ROI', _pesos(presupuesto.costo_para_roi)])
    if presupuesto.costo_m2:
        capas.append(['Costo por m² (comercial)', _pesos(presupuesto.costo_m2)])

    tabla_capas = Table(capas, colWidths=[7.5 * cm, 4 * cm], hAlign='RIGHT')
    tabla_capas.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('LINEABOVE', (0, len(capas) - 2), (-1, len(capas) - 2), 1, colors.black),
        ('FONTNAME', (0, len(capas) - 2), (-1, len(capas) - 2), 'Helvetica-Bold'),
        # La fila del ROI va destacada: es la que mueve el semáforo de la
        # oportunidad, y sin señalarla se confunde con el total comercial.
        ('TEXTCOLOR', (0, len(capas) - 1), (-1, len(capas) - 1), VERDE),
        ('FONTNAME', (0, len(capas) - 1), (-1, len(capas) - 1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    elementos.append(tabla_capas)

    nota = Paragraph(
        'El costo directo suele ser el 60–70% del total: las capas de indirectos, '
        'contingencia y utilidad no son opcionales. La utilidad es margen y no '
        'desembolso, por eso no entra en el costo que alimenta el ROI.', sub,
    )
    firma = Table([['', ''], ['Elaboró', 'Autorizó']],
                  colWidths=[7 * cm, 7 * cm], rowHeights=[1.4 * cm, 0.6 * cm])
    firma.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, GRIS),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TEXTCOLOR', (0, 1), (-1, 1), GRIS),
    ]))
    # KeepTogether: la firma no debe quedar sola en una página huérfana.
    elementos += [Spacer(1, 0.6 * cm), nota, Spacer(1, 1.2 * cm),
                  KeepTogether(firma)]

    doc.build(elementos)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

_ENCABEZADO = Font(bold=True, color='FFFFFF')
_RELLENO = PatternFill('solid', fgColor='1A7F37')
_MONEDA = '#,##0.00'


def _escribir_encabezado(hoja, titulos):
    hoja.append(titulos)
    for celda in hoja[1]:
        celda.font = _ENCABEZADO
        celda.fill = _RELLENO
        celda.alignment = Alignment(horizontal='center')


def _ajustar_anchos(hoja, anchos):
    for i, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho


def presupuesto_a_excel(presupuesto) -> bytes:
    """Dos hojas: las partidas (para trabajar sobre ellas) y el resumen de capas."""
    libro = Workbook()

    hoja = libro.active
    hoja.title = 'Partidas'
    _escribir_encabezado(hoja, ['Categoría', 'Concepto', 'Unidad', 'Cantidad', 'P.U.', 'Importe'])
    for grupo in _filas_por_categoria(presupuesto):
        for partida in grupo['partidas']:
            hoja.append([
                grupo['etiqueta'], partida.descripcion, partida.get_unidad_display(),
                float(partida.cantidad), float(partida.pu), float(partida.importe),
            ])
    for fila in hoja.iter_rows(min_row=2, min_col=4, max_col=6):
        for celda in fila:
            celda.number_format = _MONEDA
    _ajustar_anchos(hoja, [26, 52, 10, 12, 14, 16])

    resumen = libro.create_sheet('Resumen')
    _escribir_encabezado(resumen, ['Concepto', 'Importe'])
    for etiqueta, monto in _capas(presupuesto):
        resumen.append([etiqueta, float(monto)])
    resumen.append(['TOTAL (comercial)', float(presupuesto.total)])
    resumen.append(['Costo que alimenta el ROI', float(presupuesto.costo_para_roi)])
    if presupuesto.area_m2:
        resumen.append(['Área (m²)', float(presupuesto.area_m2)])
        resumen.append(['Costo por m² (comercial)', float(presupuesto.costo_m2)])
    for fila in resumen.iter_rows(min_row=2, min_col=2, max_col=2):
        for celda in fila:
            celda.number_format = _MONEDA
    resumen['A1'].alignment = Alignment(horizontal='left')
    _ajustar_anchos(resumen, [34, 18])

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def cierre_a_excel(presupuesto) -> bytes:
    """Reporte de cierre: presupuestado vs. real, órdenes de cambio y KPIs (§4.6).

    Es el documento de after-action. Se genera aunque la obra no esté cerrada:
    sirve igual como corte semanal del §4.3.
    """
    libro = Workbook()

    hoja = libro.active
    hoja.title = 'Presupuestado vs real'
    _escribir_encabezado(hoja, ['Categoría', 'Partida', 'Presupuestado', 'Real', 'Desviación', '% desv.'])
    for grupo in _filas_por_categoria(presupuesto):
        for partida in grupo['partidas']:
            hoja.append([
                grupo['etiqueta'], partida.descripcion,
                float(partida.importe), float(partida.gasto_real),
                float(partida.desviacion), float(partida.desviacion_pct),
            ])
    for fila in hoja.iter_rows(min_row=2, min_col=3, max_col=5):
        for celda in fila:
            celda.number_format = _MONEDA
    _ajustar_anchos(hoja, [26, 46, 16, 16, 16, 10])

    gastos = libro.create_sheet('Gastos')
    _escribir_encabezado(gastos, ['Fecha', 'Descripción', 'Partida', 'Proveedor', 'Importe', 'Factura'])
    for gasto in presupuesto.gastos.select_related('partida').all():
        gastos.append([
            gasto.fecha, gasto.descripcion,
            gasto.partida.descripcion if gasto.partida else 'Sin imputar',
            gasto.proveedor, float(gasto.importe_real), gasto.factura,
        ])
    for fila in gastos.iter_rows(min_row=2, min_col=5, max_col=5):
        for celda in fila:
            celda.number_format = _MONEDA
    _ajustar_anchos(gastos, [12, 46, 34, 24, 16, 20])

    ordenes = libro.create_sheet('Órdenes de cambio')
    _escribir_encabezado(ordenes, ['Fecha', 'Descripción', 'Motivo', 'Fuente', 'Importe', 'Estado', 'Resolvió'])
    for orden in presupuesto.ordenes_cambio.select_related('aprobado_por').all():
        ordenes.append([
            orden.fecha, orden.descripcion, orden.motivo,
            orden.get_fuente_display(), float(orden.importe),
            orden.get_estado_display(),
            orden.aprobado_por.username if orden.aprobado_por else '',
        ])
    for fila in ordenes.iter_rows(min_row=2, min_col=5, max_col=5):
        for celda in fila:
            celda.number_format = _MONEDA
    _ajustar_anchos(ordenes, [12, 40, 48, 30, 16, 14, 16])

    kpis = libro.create_sheet('KPIs')
    _escribir_encabezado(kpis, ['Indicador', 'Valor'])
    indicadores = [
        ('Presupuesto base (sin tocar reserva)', float(presupuesto.presupuesto_base)),
        ('Techo (base + contingencia)', float(presupuesto.base_con_contingencia)),
        ('Gasto real', float(presupuesto.gasto_real)),
        ('Desviación acumulada', float(presupuesto.desviacion_acumulada)),
        ('Desviación acumulada (%)', float(presupuesto.desviacion_acumulada_pct)),
        ('Avance de gasto sobre el techo (%)', float(presupuesto.avance_gasto_pct)),
        ('Contingencia consumida (%)', float(presupuesto.contingencia_consumida_pct)),
        ('Contingencia disponible', float(presupuesto.contingencia_disponible)),
        ('Órdenes de cambio aprobadas', presupuesto.ordenes_aprobadas_count),
        ('Monto de órdenes aprobadas', float(presupuesto.monto_ordenes_aprobadas)),
        ('Semáforo', presupuesto.semaforo_display),
    ]
    if presupuesto.costo_real_m2:
        indicadores.append(('Costo real por m²', float(presupuesto.costo_real_m2)))
    if presupuesto.costo_roi_m2:
        indicadores.append(('Costo por m² estimado (para ROI)', float(presupuesto.costo_roi_m2)))
    for etiqueta, valor in indicadores:
        kpis.append([etiqueta, valor])
    _ajustar_anchos(kpis, [38, 20])

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
